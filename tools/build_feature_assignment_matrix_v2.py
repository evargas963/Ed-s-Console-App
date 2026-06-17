"""Feature -> FULL-STACK -> horizon ASSIGNMENT MATRIX v2 (feature epic deliverable).

Operator brief (2026-06-01):
  - COMPREHENSIVE: full 7-model stack (XGB, LSTM, Transformer, Meta, Monte Carlo, Regime, Bayesian Fusion)
    across all 4 horizons (1c/5c/15c/60c), full feature universe.
  - MUST NOT LEAD CURSOR. Cursor uses independent judgement. Therefore the workbook STRICTLY
    SEPARATES three tiers:
        [FACT]     — verified from code/DB this session (file:line / query), not opinion.
        [RESEARCH] — cited external finding (lean deep-research run, 35 verified claims).
        [FRAMING]  — Claude's hypothesis/opinion. Explicitly flagged as challengeable.
  - GUARDRAILS = the standard (institutional desk / MIT professor / professional trader),
    stated up front as what should GUIDE Cursor's reasoning — NOT Claude's conclusions.
  - Cursor is asked to independently verify, agree/disagree per row, and add its own
    research + ideas in dedicated columns/sheet.

Read-only on the DB. No production path imports this. Output is an untracked operator artifact.
"""
from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from typing import Any

from governed_stack_contract import (
    ABLATION_ANCHOR_TICKERS,
    FEATURE_ABLATION_ML_STACK_LAYERS,
    FULL_STACK_MODEL_DISPLAY,
    FULL_STACK_MODEL_LAYERS,
    STACK_AUTHORITY_LAYERS,
)

HORIZONS = ["1c", "5c", "15c", "60c"]
# FULL STACK — all seven models on every ablation horizon (operator binding).
STACK = list(FULL_STACK_MODEL_DISPLAY)
FULL_STACK_LAYERS = list(FULL_STACK_MODEL_LAYERS)
# O-56: xgb/lstm/transformer layers that consume raw features — the per-model feature-ablation dimension.
# (Meta / MC / Regime / Fusion consume base-model or upstream outputs — scored in stack-authority pass.)
FEATURE_ABLATION_MODELS = list(FEATURE_ABLATION_ML_STACK_LAYERS)
# Upper stack layers scored in stack-authority pass (layer lift comparisons).
STACK_LAYERS = list(STACK_AUTHORITY_LAYERS)
FULL_STACK_ABLATION_MODES = (
    "xgb_only",
    "lstm_only",
    "transformer_only",
    "xgb_plus_lstm",
    "xgb_plus_transformer",
    "xgb_plus_lstm_plus_transformer",
    "meta_stack",
    "fusion_without_mc",
    "full_fusion",
)
# ML stack layer comparisons — "does each xgb/lstm/transformer layer earn its place in the stack?" These are
# distinct from the upper-layer (meta/MC/fusion) comparisons: a base-model lift asks whether
# adding LSTM/Transformer to the triplet improves log_loss over a leaner base set, on the live
# inference path. Kept separate from STACK_LAYER_COMPARISONS so the upper-layer contract stays
# exactly {meta, monte_carlo, fusion}.
BASE_MODEL_COMPARISONS = {
    "lstm_over_xgb": {
        "baseline": "xgb_only",
        "treatment": "xgb_plus_lstm",
        "metric": "multiclass_log_loss",
        "description": "Does LSTM add signal beyond XGB alone (0.40+0.35 renormalized)?",
    },
    "transformer_over_xgb": {
        "baseline": "xgb_only",
        "treatment": "xgb_plus_transformer",
        "metric": "multiclass_log_loss",
        "description": "Does Transformer add signal beyond XGB alone (0.40+0.25 renormalized)?",
    },
    "transformer_over_pair": {
        "baseline": "xgb_plus_lstm",
        "treatment": "xgb_plus_lstm_plus_transformer",
        "metric": "multiclass_log_loss",
        "description": "Does Transformer earn its place on top of XGB+LSTM (full triplet)?",
    },
}
STACK_LAYER_COMPARISONS = {
    "meta": {
        "baseline": "xgb_plus_lstm_plus_transformer",
        "treatment": "meta_stack",
        "metric": "multiclass_log_loss",
        "description": "Trained meta-learner vs fixed 40/35/25 weighted triplet.",
    },
    "monte_carlo": {
        "baseline": "fusion_without_mc",
        "treatment": "full_fusion",
        "metric": "multiclass_log_loss",
        "description": "MC fusion adjustment vs Bayesian fusion without MC.",
    },
    "regime": {
        "baseline": "meta_stack",
        "treatment": "fusion_without_mc",
        "metric": "multiclass_log_loss",
        "description": (
            "Volatility + market regime + rules context (signals.py vol/regime stages) on the "
            "production fusion path — fusion_without_mc vs meta_stack; MC excluded."
        ),
    },
    "fusion": {
        "baseline": "meta_stack",
        "treatment": "fusion_without_mc",
        "metric": "multiclass_log_loss",
        "description": "Bayesian fusion posterior (bayesian_fusion.fuse) vs meta stack alone.",
    },
}

HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5496")
GRP = PatternFill("solid", fgColor="D6DCE4")
FACT = PatternFill("solid", fgColor="DDEBF7")     # blue-ish = verified
RESEARCH = PatternFill("solid", fgColor="E2EFDA")  # green-ish = cited
FRAMING = PatternFill("solid", fgColor="FFF2CC")   # amber = opinion/challengeable
KEEP = PatternFill("solid", fgColor="C6EFCE")
TEST = PatternFill("solid", fgColor="FFEB9C")
ADD = PatternFill("solid", fgColor="BDD7EE")
DROP = PatternFill("solid", fgColor="FFC7CE")
NA = PatternFill("solid", fgColor="F2F2F2")
CURSORCOL = PatternFill("solid", fgColor="FCE4D6")  # orange = Cursor fills this in
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUS_FILL = {"KEEP": KEEP, "TEST": TEST, "ADD": ADD, "DROP": DROP, "n/a": NA, "": None}


def _wrapcell(ws, r, c, v, fill=None, font=None, align=WRAP, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.alignment = align
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = BORDER
    return cell


# ──────────────────────────────────────────────────────────────────────────────
# FEATURE GROUPS — same canonical set, now with the cross-horizon ladder facts
# baked in. assignment cells are FRAMING (Claude hypothesis), explicitly challengeable.
# ──────────────────────────────────────────────────────────────────────────────
def A(spec):
    out = {}
    for h in HORIZONS:
        out[h] = spec.get(h, spec.get("*", ""))
    return out


# group: key,label,members,tier-tag,source,pop,current_consumers,hz(assignment per horizon: ANY-model use), research, framing_note
GROUPS = [
    ("price_candle", "Price / candle dynamics", "candle_body_pct, candle_range_pct, body_range_ratio, cat_candle_direction",
     "Schwab leaf", "100%", "XGB, LSTM(seq), Transformer(seq)",
     A({"1c": "KEEP", "5c": "KEEP", "15c": "TEST", "60c": "DROP"}),
     "[RESEARCH] micro price action = 1-5m signal, decays long (lean-research Q1).",
     "[FRAMING] strongest at 1c/5c; question whether it adds at 60c."),
    ("vwap", "VWAP distance / side", "vwap_dist_pts, vwap_dist_pct, cat_vwap_side",
     "Derived", "100%", "XGB, LSTM(seq), Transformer(seq)",
     A({"1c": "KEEP", "5c": "KEEP", "15c": "KEEP", "60c": "TEST"}),
     "[RESEARCH] mean-reversion/acceptance anchor; intraday-swing strongest.",
     "[FRAMING] core spine feature, all horizons."),
    ("net_gamma", "Net gamma + sign (+ ΔGEX candidate)", "net_gamma, gamma_positive, m5_net_gamma",
     "Schwab leaf", "100%", "XGB, LSTM(seq subset)",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] dealer short-gamma -> intraday momentum; ΔGEX (first diff) > level OOS (Q1). SLOW: weak at 1c.",
     "[FRAMING] add ΔGEX as a cheap first-difference; sign is tree-friendly."),
    ("net_delta", "Net delta + sign", "net_delta, delta_positive, m5_net_delta",
     "Schwab leaf", "100%", "XGB, LSTM(seq subset)",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] DEX = directional dealer positioning, slower horizon.", "[FRAMING] not a 1c signal."),
    ("charm", "Charm (delta decay)", "charm_net, charm_delta_agree, cat_charm_direction, cat_charm_magnitude, m5_charm_net",
     "Derived (2nd-order, NO_SCHWAB charm primitive)", "97-99%", "XGB, LSTM(seq subset)",
     A({"1c": "DROP", "5c": "TEST", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] end-of-day delta-decay pin pressure -> close/60c/OPEX.",
     "[FRAMING] 60c-side feature; governed derivation — verify formula."),
    ("gamma_walls", "Gamma walls + inflection", "dist_call/put_gamma_wall_pct, dist_gamma_inflection_pct (+m5)",
     "Schwab leaf", "99-100%", "XGB, LSTM(seq subset)",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "TEST"}),
     "[RESEARCH] pin/acceleration trigger; threshold crossings = tree-native.",
     "[FRAMING] core LFE Layer-2 structure too."),
    ("delta_walls", "Delta walls + inflection", "dist_call/put_delta_wall_pct, dist_delta_inflection_pct (+m5)",
     "Schwab leaf", "99-100%", "XGB",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] correlated w/ gamma walls -> grouped permutation.",
     "[FRAMING] XGB-only today; candidate ADD to nets if ablation lifts."),
    ("oi_walls", "OI walls + put/call ratio", "dist_call/put_oi_wall_pct, put_call_oi_ratio (+m5)",
     "Schwab leaf", "100%", "XGB",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] OI concentration = sticky/slow; PCR regime proxy. OI-migration (day/day) candidate.",
     "[FRAMING] slow structure; longer horizons."),
    ("vanna_walls", "Vanna walls", "dist_call/put_vanna_wall_pct (+m5)",
     "Derived (2nd-order, NO_SCHWAB)", "100%", "XGB",
     A({"1c": "DROP", "5c": "TEST", "15c": "TEST", "60c": "TEST"}),
     "[RESEARCH] vol-of-delta dealer flow; lower-confidence.",
     "[FRAMING] first ablation chopping-block candidate."),
    ("pin", "Pin width / score", "pin_width_pct, pin_score (+m5)",
     "Derived", "99%", "XGB",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] pinning intensifies into expiry/close.", "[FRAMING] 15-60c."),
    ("zone", "Zone / structure state", "cat_zone, cat_prev_zone, zone_since_bars, breakout_score",
     "Derived (key-levels)", "100%", "XGB(cat+num), LSTM(zone in seq)",
     A({"1c": "KEEP", "5c": "KEEP", "15c": "KEEP", "60c": "TEST"}),
     "[RESEARCH] acceptance/rejection structure -> feeds LFE.", "[FRAMING] spine feature."),
    ("iv", "Implied vol level / rank / direction", "iv_level, iv_rank, cat_iv_direction (+m5)",
     "Schwab leaf", "96-98%", "XGB, LSTM(seq: iv_level)",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] vol regime is slow -> longer horizons; iv_rank = regime-normalized.", "[FRAMING] 5-60c."),
    ("vix", "VIX level / change / bucket", "vix_level, vix_vs_prev, cat_vix_bucket",
     "Schwab leaf ($VIX, panel_auto)", "100%", "XGB, LSTM(seq: vix_level)",
     A({"1c": "DROP", "5c": "TEST", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] macro vol regime; near-static intraday.", "[FRAMING] noise at 1c."),
    ("index_spreads", "Cross-index spreads (Slice A)", "qqq_vs_spy, spy_iwm_divergence",
     "Schwab leaf", "100%", "XGB",
     A({"*": "KEEP"}),
     "[RESEARCH] precomputed spread vs train-time align — ablation decides delta.",
     "[FRAMING] keep separate from index_xasset align/chg families (override no-collapse)."),
    ("rates", "Rates (TNX)", "tnx_yield, tnx_chg",
     "Schwab leaf", "100%", "XGB",
     A({"1c": "TEST", "5c": "KEEP", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] rates -> growth/financials regime for SPY.",
     "[FRAMING] Slice A free cols."),
    ("index_xasset", "Index cross-asset (SPY/QQQ/IWM)", "spy/qqq/iwm_chg_pct, *_weighted_push, *_align, cross_avg/std_chg (+m5)",
     "Schwab leaf", "100%", "XGB, LSTM(seq)",
     A({"*": "KEEP"}),
     "[RESEARCH] beta/regime context all horizons. Lead-lag (leaders->SPY) FRAGILE OOS — do NOT over-invest.",
     "[FRAMING] spine; watch redundancy via grouped permutation."),
    ("megacap", "Mega-cap constituents", "nvda/aapl/msft/amzn/googl/avgo/meta/tsla_chg_pct",
     "Schwab leaf", "100%", "XGB",
     A({"1c": "TEST", "5c": "TEST", "15c": "TEST", "60c": "DROP"}),
     "[RESEARCH] basket DRIVES SPY/QQQ (operator). 8 corr features -> grouped permutation.",
     "[FRAMING] ADD candidate as a constituent-trajectory stream for nets."),
    ("breadth_etf", "Sector / breadth ETFs", "kre/xbi/psci/xrt_chg_pct",
     "Schwab leaf (sector ETFs, panel_auto)", "100%", "XGB",
     A({"1c": "DROP", "5c": "TEST", "15c": "TEST", "60c": "KEEP"}),
     "[RESEARCH] breadth = regime/macro -> IWM especially.", "[FRAMING] most relevant to IWM model."),
    ("order_flow", "Order-flow imbalance (OFI / L1)", "bid_ask_imbalance (NEVER WIRED), flow_imbalance (live), imbalance_buy/sell_pressure",
     "Schwab L1 — DERIVABLE", "see [FACT] note", "XGB (via flow_imbalance fallback)",
     A({"1c": "ADD", "5c": "ADD", "15c": "TEST", "60c": "DROP"}),
     "[RESEARCH] single-level OFI = DOMINANT 1-5m signal (65-87% of mid-price variance); 1c=collapse horizon (Q1).",
     "[FACT] `bid_ask_imbalance` column does NOT EXIST in snapshots or snapshots_1m_normalized (verified query); code falls back to `flow_imbalance` (present, ~100% SPY). [FRAMING] THE #1 feature gap — build proper single-level OFI from L1."),
    ("volume", "Volume", "candle_volume_log, volume_ratio",
     "Schwab leaf", "100%", "XGB",
     A({"1c": "KEEP", "5c": "KEEP", "15c": "TEST", "60c": "DROP"}),
     "[RESEARCH] volume burst = short-horizon conviction; volume-clock for triple-barrier.",
     "[FRAMING] ADD to LSTM/Transformer (code comment already lists it as planned)."),
    ("microstructure", "Microstructure derived", "spread, smart_money_score, absorption_score, continuation_score",
     "Derived (spread=Schwab bid/ask; scores=composites)", "spread 100% / absorption+continuation 73-85%", "XGB",
     A({"1c": "TEST", "5c": "TEST", "15c": "TEST", "60c": "DROP"}),
     "[RESEARCH] spread=liquidity/cost; behavioral scores feed LFE acceptance layer.",
     "[FACT] absorption/continuation thinner pop (73-85%, lower on IWM). [FRAMING] verify provenance."),
    ("regime_composites", "Regime / signal composites", "cat_combined_signal, cat_combined_conviction (EXCLUDE), cat_pressure_label, cat_pressure_trend, cat_liquidity_behavior_label, cat_session_bucket",
     "Derived (dealer flow + session clock)", "100%", "XGB(cat)",
     A({"1c": "TEST", "5c": "TEST", "15c": "KEEP", "60c": "KEEP"}),
     "[RESEARCH] categorical regime = tree-native.",
     "[FACT] combined_signal/combined_conviction = ms.call_signal/call_conviction AFTER compute_call (server.py:4893-4894) = confirmed circular leakage; feature_contracts forbidden family; curation gate EXCLUDED. [FACT] pressure_label/trend = DPI/hedging-flow direction BEFORE snapshot write (server.py:4687-4695) = pre-decision KEEP. [FRAMING] split the row: ablation uses pressure/liquidity/session only."),
    ("time", "Time-of-day", "time_sin, time_cos, time_progress, minutes_since_open, et_hour, et_minute, cat_session_bucket",
     "Derived (ts_utc -> ET clock)", "100%", "XGB (all); nets implicit via seq position",
     A({"*": "KEEP"}),
     "[RESEARCH] intraday seasonality (open/close), strongest 60c/open.",
     "[FRAMING] keep on XGB; test necessity on nets (seq position may make it redundant)."),
]

# Cursor independent scrutiny @ 2026-06-01 (re-verified code/DB; fills Feature Master cols 12-13).
CURSOR_VERDICTS: dict[str, tuple[str, str]] = {
    "price_candle": ("Agree", "1c/5c KEEP, 60c DROP — micro candle shape decays; matches curation iv_level>iv_rank pattern."),
    "vwap": ("Agree", "Spine all horizons; 60c TEST reasonable. spot/vix must NOT cluster (override)."),
    "net_gamma": ("Agree", "ΔGEX P0 ADD; 1c TEST not KEEP — slow dealer positioning."),
    "net_delta": ("Agree", "Same ladder as gamma; grouped MDA with gamma family."),
    "charm": ("Agree", "1c DROP, 60c KEEP — EOD decay story holds."),
    "gamma_walls": ("Agree", "Core structure; vanna rep-over-gamma OK for ablation WITH override note."),
    "delta_walls": ("Agree", "XGB-only today; nets ADD candidate post-ablation."),
    "oi_walls": ("Agree", "Slow structure; OI-migration P2 candidate."),
    "vanna_walls": ("Agree", "First chop candidate; TEST not KEEP at 5c+."),
    "pin": ("Agree", "15-60c pin story; keep in spine."),
    "zone": ("Agree", "Spine; breakout_score feeds LFE."),
    "iv": ("Agree", "Rep iv_level over iv_rank (curation + override). 5-60c primary."),
    "vix": ("Agree", "1c DROP; never collapse vix_level into spot (override)."),
    "index_xasset": ("Agree w/ override", "KEEP all horizons BUT Spearman must NOT pick one chg_pct rep — spy/qqq/iwm + weighted_push stay separate (feature_curation_overrides.json)."),
    "megacap": ("Agree", "Grouped permutation mandatory; 60c DROP for index anchors; constituent stream for nets = TEST."),
    "breadth_etf": ("Agree", "IWM-weighted; 1c DROP."),
    "order_flow": ("Agree", "[FACT] bid_ask_imbalance absent from normalized table; flow_imbalance 100% SPY; P0 single-level OFI wiring."),
    "volume": ("Agree", "1c/5c KEEP; ADD volume to LSTM/Transformer per lstm_data.py comment."),
    "microstructure": ("Agree", "spread KEEP short horizons; absorption/continuation thinner on IWM — population gate catches."),
    "regime_composites": ("Agree w/ split", "combined_* EXCLUDE (leakage closed); pressure/trend + liquidity_behavior + session_bucket KEEP (pre-decision)."),
    "time": ("Mostly agree", "XGB KEEP all horizons; nets: TEST — seq position may subsume et_hour/et_minute."),
}

CURSOR_MC_STACK = (
    "Agree. Underutilized not unused: sizing (call_engine), narrative, Key Levels display; "
    "fusion posterior weight=0; NOT in engineer_features cone. BAR_MINUTES=5 blocks efe/eae-as-ML-feature "
    "until wall-clock alignment fix. P1 after alignment."
)

CURSOR_ARCHITECTURE = (
    "Agree (champion/challenger). Keep 12-independent stack as baseline from Stage-2 retrain; "
    "joint shared-encoder is CHALLENGER on held-out per-horizon MCC only. XGB stays independent. "
    "Cross-horizon ladder (12-group spine, not 13 — see sheet 4 correction) buys emergent coherence; "
    "enforced coherence not worth negative-transfer risk until joint beats baseline on our 1m substrate."
)

CURSOR_LABELING = (
    "Agree (method + D2). Curate-then-ablate ONE pass matches executed feature_curation_gate.py; "
    "harness MUST pin cluster reps from feature_curation_overrides.json (never auto-rep from last gate run). "
    "Triple-barrier = challenger label; dual-label backtest resolves operator 'missed move' concern before switch."
)

CURSOR_LADDER = (
    "Agree (emergent coherence thesis). [CORRECTION] Recomputed from GROUPS @ 2026-06-01: "
    "1c=17 groups, 5c/15c=21, 60c=16; spine ALL FOUR = 12 groups (not 13 — price_candle drops at 60c); "
    "1c&60c endpoints share 12 (not 13); 15c&60c share 16/21 (not 17/21). "
    "FAST-only @1c not 60c: order_flow, volume, microstructure, megacap, price_candle."
)


LEGACY_COMPOUND_MANIFEST_PATH = Path("governance/artifacts/feature_ablation_manifest.json")
EXPANDED_ABLATION_MANIFEST_PATH = Path("governance/artifacts/feature_ablation_manifest_leaf.json")
FEATURE_ABLATION_UNIVERSE_XLSX = Path("feature_ablation_universe.xlsx")
SCHWAB_DICTIONARY_PATH = Path("schwab_field_inventory/schwab_field_dictionary.csv")
SCHWAB_ABLATION_REGISTRY_PATH = Path("governance/artifacts/schwab_ablation_field_registry.json")
SNAPSHOT_CULL_LEDGER_PATH = Path("governance/artifacts/snapshot_column_cull_ledger.json")
# Operator binding: ablation pool must be ≥ this factor × registered ML cone — ablation picks winners.
MIN_ABLATION_EXPANSION_FACTOR = 2.0
ABLATION_FEATURE_GRAIN = "schwab_expanded_atomic"
ABLATION_PRIMARY_PASS = "per_model_per_horizon_atomic_permutation_importance"
ABLATION_CONFIRM_PASS = "per_model_per_horizon_atomic_drop_column_refit_on_survivors"

# Schwab dictionary categorization (research tiers — NOT winner-picking; ablation decides value).
_SCHWAB_INCLUDE_CATEGORIES = frozenset(
    {
        "greeks",
        "volatility",
        "bid_ask",
        "volume",
        "key_levels",
        "streaming_quote",
        "streaming_book",
        "options_chain",
        "time",
        "price_ohlc",
    }
)
_SCHWAB_INCLUDE_LIKELY_USE = frozenset(
    {
        "prediction_model",
        "direct_order_flow",
        "order_flow_proxy",
        "risk_model",
        "regime_detection",
        "key_levels",
    }
)
_SCHWAB_TRADE_ENDPOINTS = frozenset({"quotes", "chains", "pricehistory", "streaming", "movers", "instruments"})

def _camel_to_snake(name: str) -> str:
    import re

    s1 = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    return re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s1).lower()


def _schwab_primary_endpoint(source_endpoints: str) -> str:
    return (source_endpoints or "").split(";")[0].strip()


def categorize_schwab_field(row: dict) -> str:
    """Catalog tier for one Schwab dictionary row — categorization only, not ablation verdict."""
    category = str(row.get("category") or "unknown").strip()
    likely_use = str(row.get("likely_use") or "unknown").strip()
    priority = str(row.get("priority") or "medium").strip()
    endpoint = _schwab_primary_endpoint(str(row.get("source_endpoints") or ""))

    if likely_use == "ui_only":
        return "EXCLUDE_UI"
    if endpoint == "market_hours" and category == "unknown" and likely_use == "unknown":
        return "EXCLUDE_SESSION"
    if likely_use in _SCHWAB_INCLUDE_LIKELY_USE or category in _SCHWAB_INCLUDE_CATEGORIES:
        return "ML_ABLATION_CANDIDATE"
    if endpoint in _SCHWAB_TRADE_ENDPOINTS and priority in ("high", "medium") and category != "unknown":
        return "ML_INGEST_CANDIDATE"
    if endpoint in _SCHWAB_TRADE_ENDPOINTS and priority == "high":
        return "ML_INGEST_CANDIDATE"
    return "CATALOG_ONLY"


def _canonical_to_column_hint(canonical_field: str, example_raw: str = "") -> str:
    leaf = str(canonical_field or "").split(".")[-1]
    if leaf in ("*", "callExpDateMap", "putExpDateMap"):
        leaf = example_raw.split(".")[-1] if example_raw else leaf
    hint = _camel_to_snake(leaf)
    if hint in ("last_price", "lastprice"):
        return "spot"
    if hint == "volatility":
        return "iv_level"
    return hint


def _build_schwab_ablation_field_registry_payload(*, stable_time: bool = False) -> dict:
    """Build Schwab ablation registry dict from dictionary CSV (no disk I/O)."""
    import csv
    from datetime import datetime, timezone

    if not SCHWAB_DICTIONARY_PATH.is_file():
        raise FileNotFoundError(SCHWAB_DICTIONARY_PATH)
    rows_in = list(csv.DictReader(SCHWAB_DICTIONARY_PATH.open(encoding="utf-8")))
    fields: list[dict] = []
    tier_counts: dict[str, int] = {}
    for row in rows_in:
        tier = categorize_schwab_field(row)
        tier_counts[tier] = tier_counts.get(tier, 0) + 1
        fields.append(
            {
                "canonical_field": row.get("canonical_field"),
                "catalog_tier": tier,
                "category": row.get("category"),
                "likely_use": row.get("likely_use"),
                "priority": row.get("priority"),
                "source_endpoints": row.get("source_endpoints"),
                "column_hint": _canonical_to_column_hint(
                    str(row.get("canonical_field") or ""),
                    str(row.get("example_raw_field") or ""),
                ),
            }
        )
    return {
        "schema_version": "1",
        "generated_at": "stable" if stable_time else datetime.now(timezone.utc).isoformat(),
        "schwab_dictionary_path": str(SCHWAB_DICTIONARY_PATH),
        "schwab_field_count": len(fields),
        "tier_counts": tier_counts,
        "min_ablation_expansion_factor": MIN_ABLATION_EXPANSION_FACTOR,
        "fields": fields,
    }


def load_schwab_ablation_field_registry(*, write: bool = False, stable_time: bool = False) -> dict:
    """Load on-disk Schwab ablation registry, or build in memory when missing or write=True."""
    if not write and SCHWAB_ABLATION_REGISTRY_PATH.is_file():
        try:
            on_disk = json.loads(SCHWAB_ABLATION_REGISTRY_PATH.read_text(encoding="utf-8"))
            if int(on_disk.get("schwab_field_count") or 0) >= 2300 and on_disk.get("fields"):
                return on_disk
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            pass
    out = _build_schwab_ablation_field_registry_payload(stable_time=stable_time)
    if write:
        SCHWAB_ABLATION_REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
        SCHWAB_ABLATION_REGISTRY_PATH.write_text(json.dumps(out, indent=2), encoding="utf-8")
    return out


def build_schwab_ablation_field_registry(*, write: bool = True, stable_time: bool = False) -> dict:
    """Categorize all Schwab dictionary rows; write governance/artifacts/schwab_ablation_field_registry.json."""
    return load_schwab_ablation_field_registry(write=write, stable_time=stable_time)


def _registered_ml_columns() -> dict[str, set[str]]:
    """Live ingest cones: xgb tabular (incl. cf_*); LSTM streams + X_conf (cf_* not duplicated on streams)."""
    import lstm_data as l
    from ml_train import tabular_training_feature_names

    tabular = set(tabular_training_feature_names())
    conf = set(l.CONFLUENCE_FEATURES)
    sequence = tabular - conf
    lstm_ingest = sequence | conf
    return {
        "xgb": tabular,
        "lstm_5m": lstm_ingest,
        "lstm_1m": lstm_ingest,
    }


def _snapshot_expansion_columns(registered: dict[str, set[str]]) -> list[dict]:
    """Snapshot columns with trade consumers not yet in the registered cone."""
    if not SNAPSHOT_CULL_LEDGER_PATH.is_file():
        return []
    try:
        ledger = json.loads(SNAPSHOT_CULL_LEDGER_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, ValueError):
        return []
    known = set().union(*registered.values())
    out: list[dict] = []
    for col in ledger.get("columns") or []:
        if not isinstance(col, dict):
            continue
        name = str(col.get("column") or "").strip()
        verdict = str(col.get("verdict") or "")
        if not name or name in known:
            continue
        if verdict not in ("KEEP", "KEEP_LIVE", "WIRED_PENDING_DATA"):
            continue
        null_pct = float(col.get("null_pct") or 100.0)
        if null_pct >= 99.0:
            continue
        out.append(
            {
                "column": name,
                "verdict": verdict,
                "null_pct": null_pct,
                "catalog_tier": "SNAPSHOT_EXPANSION",
                "ingest_status": "in_snapshot",
            }
        )
    return sorted(out, key=lambda r: (r["null_pct"], r["column"]))


def _atomic_ablation_group(
    *,
    group_id: str,
    label: str,
    column: str,
    catalog_tier: str,
    ingest_status: str,
    schwab_lineage: list[str],
    disposition: str = "ABLATE",
) -> dict:
    """Pure feature entry — no model placement (ZERO-BIAS: ablation survivors decide model×horizon)."""
    return {
        "group_id": group_id,
        "label": label,
        "atomic_column": column,
        "disposition": disposition,
        "catalog_tier": catalog_tier,
        "ingest_status": ingest_status,
        "schwab_lineage": schwab_lineage,
    }


def per_model_ablation_cell_target(manifest: dict) -> int:
    method = manifest.get("ablation_method") or {}
    anchors = len(method.get("anchors") or list(ABLATION_ANCHOR_TICKERS))
    models = len(method.get("models") or FEATURE_ABLATION_MODELS)
    hz = len(method.get("horizons") or HORIZONS)
    ablate = len([g for g in manifest.get("groups") or [] if g.get("disposition") == "ABLATE"])
    return anchors * models * hz * ablate


def _atomic_members_from_ingest_cone(
    column: str, registered: dict[str, set[str]]
) -> tuple[list[str], list[str], list[str]]:
    """Ingest-capability metadata only — NOT placement. O-56 survivors assign model×horizon winners."""
    xgb = [column] if column in registered.get("xgb", set()) else []
    l5 = [column] if column in registered.get("lstm_5m", set()) else []
    l1 = [column] if column in registered.get("lstm_1m", set()) else []
    return xgb, l5, l1


def atomic_column_for_manifest_group(group: dict) -> str | None:
    """Single atomic column identity from a manifest group (no model stamps)."""
    col = group.get("atomic_column")
    if col:
        return str(col).strip() or None
    gid = str(group.get("group_id") or "")
    for prefix in ("reg__atomic__", "schwab__", "snap__"):
        if gid.startswith(prefix):
            return gid[len(prefix) :]
    return None


def expected_ingest_members_for_atomic_group(
    group: dict, registered: dict[str, set[str]] | None = None
) -> dict[str, list[str]]:
    """Re-derive ingest-capability from live code cone — audit-time only, not manifest fields."""
    if registered is None:
        registered = _registered_ml_columns()
    col = atomic_column_for_manifest_group(group)
    if not col:
        return {"xgb": [], "lstm_5m": [], "lstm_1m": []}
    xgb, l5, l1 = _atomic_members_from_ingest_cone(col, registered)
    return {"xgb": xgb, "lstm_5m": l5, "lstm_1m": l1}


def resolve_expanded_schwab_ablation_universe(*, write_registry: bool = False) -> dict[str, Any]:
    """Schwab-catalog-first ablation universe: categorize all 2393 leaves, expand pool ≥2× registered cone."""
    registry = load_schwab_ablation_field_registry(write=write_registry)
    registered = _registered_ml_columns()
    reg_union = set().union(*registered.values())
    reg_count = len(reg_union)

    groups: list[dict] = []
    seen_group_ids: set[str] = set()
    lineage_by_column: dict[str, list[str]] = {}

    for col in sorted(reg_union):
        gid = f"reg__atomic__{col}"
        if gid in seen_group_ids:
            continue
        seen_group_ids.add(gid)
        if col in set(getattr(__import__("lstm_data", fromlist=["CONFLUENCE_FEATURES"]), "CONFLUENCE_FEATURES", [])):
            tier = "REGISTERED_CONFLUENCE"
        else:
            tier = "REGISTERED_UNIVERSE"
        groups.append(
            _atomic_ablation_group(
                group_id=gid,
                label=f"atomic / {col}",
                column=col,
                catalog_tier=tier,
                ingest_status="in_cone",
                schwab_lineage=[],
            )
        )

    for snap in _snapshot_expansion_columns(registered):
        col = snap["column"]
        gid = f"snap__{col}"
        if gid in seen_group_ids:
            continue
        seen_group_ids.add(gid)
        groups.append(
            _atomic_ablation_group(
                group_id=gid,
                label=f"snapshot expansion / {col}",
                column=col,
                catalog_tier=snap["catalog_tier"],
                ingest_status=snap["ingest_status"],
                schwab_lineage=[],
            )
        )

    # Price-action snapshot columns (operator 2026-06-11): persisted bar-derived
    # primitives. Serving-cone registration is gated behind retrain ([REAL-GATE:
    # training-skew] PA-CONE-V8-RETRAIN), but ZERO-BIAS requires them in the
    # ablation universe NOW — DB-wire reconcile flips them in_cone for scoring.
    from features.signal_layer_v1 import SNAPSHOT_PRICE_ACTION_COLUMNS

    for col, _layer_key in SNAPSHOT_PRICE_ACTION_COLUMNS:
        gid = f"snap__{col}"
        if gid in seen_group_ids or col in reg_union:
            continue
        seen_group_ids.add(gid)
        groups.append(
            _atomic_ablation_group(
                group_id=gid,
                label=f"price-action snapshot / {col}",
                column=col,
                catalog_tier="SNAPSHOT_EXPANSION",
                ingest_status="in_snapshot",
                schwab_lineage=[],
            )
        )

    pool_fields = [
        f
        for f in registry["fields"]
        if f.get("catalog_tier") in ("ML_ABLATION_CANDIDATE", "ML_INGEST_CANDIDATE")
    ]
    for field in pool_fields:
        hint = str(field.get("column_hint") or "").strip()
        canonical = str(field.get("canonical_field") or "")
        if not hint:
            continue
        lineage_by_column.setdefault(hint, []).append(canonical)

    min_ablate_groups = int(math.ceil(reg_count * MIN_ABLATION_EXPANSION_FACTOR))
    for hint, lineages in sorted(lineage_by_column.items()):
        if hint in reg_union:
            for g in groups:
                if g.get("atomic_column") == hint:
                    existing = g.setdefault("schwab_lineage", [])
                    for leaf in lineages:
                        if leaf not in existing:
                            existing.append(leaf)
            continue
        gid = f"schwab__{hint}"
        if gid in seen_group_ids:
            continue
        seen_group_ids.add(gid)
        tier = "ML_INGEST_CANDIDATE"
        for f in pool_fields:
            if f.get("column_hint") == hint:
                tier = str(f.get("catalog_tier") or tier)
                break
        groups.append(
            _atomic_ablation_group(
                group_id=gid,
                label=f"schwab catalog / {hint}",
                column=hint,
                catalog_tier=tier,
                ingest_status="not_wired",
                schwab_lineage=sorted(set(lineages)),
            )
        )

    ablate_groups = [g for g in groups if g.get("disposition") == "ABLATE"]
    if len(ablate_groups) < min_ablate_groups:
        raise RuntimeError(
            f"Expanded ablation universe too small: {len(ablate_groups)} groups "
            f"< {min_ablate_groups} (= {MIN_ABLATION_EXPANSION_FACTOR}× registered cone {reg_count})"
        )

    anchors = list(ABLATION_ANCHOR_TICKERS)
    per_model_cells = len(anchors) * len(FEATURE_ABLATION_MODELS) * len(HORIZONS) * len(ablate_groups)
    payload = {
        "schema_version": "4_schwab_expanded",
        "source": "tools/build_feature_assignment_matrix_v2.py::resolve_expanded_schwab_ablation_universe",
        "ablation_method": {
            "primary_pass": ABLATION_PRIMARY_PASS,
            "primary_pass_description": (
                "Per ML stack layer (xgb/lstm/transformer) × horizon (1c/5c/15c/60c): train on chronological "
                "holdout and permute/drop ONE atomic feature at a time; measure that model's MCC delta. "
                "Survivors resolve per (model, horizon) cell only."
            ),
            "feature_grain": ABLATION_FEATURE_GRAIN,
            "universe_policy": (
                "Schwab dictionary categorization (all 2393 rows) + registered ML cone + snapshot "
                f"expansion; pool size ≥ {MIN_ABLATION_EXPANSION_FACTOR}× registered cone; ablation "
                "decides winners — agents MUST NOT pre-pick stack-consumed-only subsets."
            ),
            "schwab_registry_path": str(SCHWAB_ABLATION_REGISTRY_PATH),
            "decision_mode": "per_model_holdout",
            "decision_metric": "mcc_delta",
            "models": list(FEATURE_ABLATION_MODELS),
            "full_stack_layers": list(FULL_STACK_LAYERS),
            "horizons_required": list(HORIZONS),
            "horizons_binding": "all_four_mandatory__partial_horizon_runs_rejected",
            "confirm_pass": ABLATION_CONFIRM_PASS,
            "confirm_pass_description": (
                "After primary atomic permutation: per model × horizon, refit on survivors-only "
                "(dropped atomic features removed) and confirm held-out MCC is not worse."
            ),
            "confirm_pass_cli": "python tools/feature_curation_gate.py --ablation-confirm",
            "anchors": anchors,
            "pool_tickers": anchors,
            "stage3_scoring_mode": "pooled_ticker_rows_one_log_loss",
            "stage3_grid": "captured_cone_feature_x_horizon",
            "horizons": list(HORIZONS),
        },
        "groups": groups,
        "totals": {
            "schwab_dictionary_rows": registry["schwab_field_count"],
            "schwab_tier_counts": registry["tier_counts"],
            "registered_ml_cone_columns": reg_count,
            "min_ablation_expansion_factor": MIN_ABLATION_EXPANSION_FACTOR,
            "ablation_group_count": len(ablate_groups),
            "per_model_feature_cell_count": per_model_cells,
            "expansion_ratio_vs_cone": round(len(ablate_groups) / max(reg_count, 1), 2),
        },
    }
    return payload


def load_ablation_cell_target(manifest_path: Path | None = None) -> int:
    """Primary matrix cell target from on-disk expanded manifest (0 if missing)."""
    mp = manifest_path or EXPANDED_ABLATION_MANIFEST_PATH
    if not mp.is_file():
        return 0
    try:
        data = json.loads(mp.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, ValueError):
        return 0
    return per_model_ablation_cell_target(data)


def resolve_ablation_universe(*, write_registry: bool = False) -> dict[str, Any]:
    """Canonical ablation manifest — Schwab-catalog expanded universe (NOT stack-consumed-only)."""
    return resolve_expanded_schwab_ablation_universe(write_registry=write_registry)


def write_feature_ablation_manifest(path: Path | None = None) -> Path:
    out_path = path or EXPANDED_ABLATION_MANIFEST_PATH
    payload = resolve_ablation_universe(write_registry=True)
    try:
        from db import DB_PATH
        from tools.feature_curation_gate import reconcile_manifest_ingest_status_to_db_wire

        if Path(DB_PATH).is_file():
            payload = reconcile_manifest_ingest_status_to_db_wire(payload, str(DB_PATH))
    except Exception:
        pass
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    write_feature_ablation_universe_xlsx(payload=payload)
    return out_path


def _join_member_list(items) -> str:
    if isinstance(items, list):
        return ", ".join(str(x) for x in items)
    return str(items) if items else ""


def write_feature_ablation_universe_xlsx(
    *,
    payload: dict[str, Any] | None = None,
    path: Path | None = None,
) -> Path:
    """Single operator workbook — one sheet, ablation groups only (machine JSON stays separate)."""
    data = payload if payload is not None else resolve_ablation_universe()
    groups = [g for g in (data.get("groups") or []) if g.get("disposition") == "ABLATE"]
    totals = data.get("totals") or {}
    out_path = path or FEATURE_ABLATION_UNIVERSE_XLSX

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Ablation List"
    headers = [
        "group_id",
        "label",
        "atomic_column",
        "catalog_tier",
        "ingest_status",
        "schwab_lineage",
    ]
    for c, name in enumerate(headers, 1):
        _wrapcell(ws, 1, c, name, fill=HDR, font=WHITE)
    widths = [28, 36, 24, 20, 14, 40]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    for r, g in enumerate(groups, 2):
        ws.cell(r, 1, g.get("group_id", ""))
        ws.cell(r, 2, g.get("label", ""))
        ws.cell(r, 3, g.get("atomic_column", ""))
        ws.cell(r, 4, g.get("catalog_tier", ""))
        ws.cell(r, 5, g.get("ingest_status", ""))
        ws.cell(r, 6, _join_member_list(g.get("schwab_lineage", [])))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(groups) + 1)}"
    wb.properties.title = "Feature Ablation Universe"
    wb.properties.subject = (
        f"{len(groups)} ablation groups; "
        f"expansion {totals.get('expansion_ratio_vs_cone')}× cone; "
        f"grain {ABLATION_FEATURE_GRAIN}"
    )
    wb.save(out_path)
    return out_path


def _ladder_stats() -> dict[str, Any]:
    """Recompute cross-horizon overlap from GROUPS (single source for ladder sheet)."""

    def groups_at_hz(h: str) -> set[str]:
        out: set[str] = set()
        for g in GROUPS:
            key, _label, _members, _src, _pop, _cons, hz, _research, _framing = g
            st = hz[h]
            if st and st != "DROP":
                out.add(key)
        return out

    by_h = {h: groups_at_hz(h) for h in HORIZONS}
    pairs = [("1c", "5c"), ("5c", "15c"), ("15c", "60c"), ("1c", "60c")]
    pair_txt = []
    for a, b in pairs:
        shared = by_h[a] & by_h[b]
        pair_txt.append(f"  {a} & {b}: {len(shared)} shared ({len(shared)}/{len(by_h[a])} of {a}; {len(shared)}/{len(by_h[b])} of {b})")
    spine = set.intersection(*by_h.values())
    fast = by_h["1c"] - by_h["60c"]
    slow = by_h["60c"] - by_h["1c"]
    return {
        "by_h": by_h,
        "pair_txt": pair_txt,
        "spine": sorted(spine),
        "fast": sorted(fast),
        "slow": sorted(slow),
    }


def build():
    wb = openpyxl.Workbook()

    # ===== Sheet 1: HOW TO READ THIS / CURSOR INSTRUCTIONS =====
    ws = wb.active
    ws.title = "READ ME — Cursor brief"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 122
    lines = [
        ("title", "FEATURE -> FULL-STACK -> HORIZON ASSIGNMENT MATRIX  (v2 DRAFT)"),
        ("sub", "Built 2026-06-01 @ tip f73365c. Cursor scrutiny pass 2026-06-01 — Feature Master cols 12-13 filled; ladder recomputed from GROUPS."),
        ("blank", ""),
        ("h", "TO CURSOR — how to use this document"),
        ("p", "1. This is Claude's FRAMING, not a conclusion to ratify. You are asked to INDEPENDENTLY VERIFY every [FACT], re-derive the horizon assignments with your OWN judgement, and AGREE or DISAGREE per row — adding your reasoning + any research YOU run."),
        ("p", "2. Three tiers are colour-separated so you can tell evidence from opinion:"),
        ("fact", "   [FACT]  (blue)  — verified from code/DB this session (query or file:line). Re-verify; flag any you cannot reproduce."),
        ("research", "   [RESEARCH] (green) — cited external finding (lean deep-research run; sources in 'Research basis' sheet). Challenge or extend with your own sources."),
        ("framing", "   [FRAMING] (amber) — Claude's hypothesis/opinion. EXPLICITLY challengeable. Do not treat as settled."),
        ("p", "3. Fill the CURSOR columns (orange) on every sheet: Agree?/Disagree + your reasoning + your sources. Where you disagree, propose the alternative."),
        ("p", "4. The horizon-architecture question (independent vs joint) is a STACK decision, bigger than this matrix — give it your own verdict on the 'Architecture' sheet."),
        ("blank", ""),
        ("h", "GUARDRAILS — the standard that should guide YOUR reasoning (not Claude's conclusions)"),
        ("p", "Every judgement here is held to: (a) would an MIT professor accept the method (correctness, held-out validation, no hand-waving)? (b) would a professional/institutional trading desk run it (tradeable edge, honest measurement, fail-closed on no-edge)? (c) is it world-class, not 'good enough'? If a row fails the standard, say so — even if Claude proposed it."),
        ("p", "Hard rules in force: Schwab-first (leaf or governed NO_SCHWAB exception, no fabricated values); no leakage; held-out / purged-CV evaluation; MCC + per-class recall over accuracy; no carried residuals."),
        ("blank", ""),
        ("h", "WHAT CHANGED FROM v1 (corrections you should check)"),
        ("p", "  - FULL 6-LAYER STACK now shown (XGB, LSTM, Transformer, Meta, Monte Carlo, Regime/Rules-Fusion), all 4 horizons — operator: 'use the entire stack; not doing so is cheating ourselves.' v1 wrongly showed only xgb/lstm/transformer-only framing."),
        ("p", "  - CROSS-HORIZON SHARED-FEATURE LADDER added (operator insight: coherence emerges from shared inputs up the ladder). See 'Cross-horizon ladder' sheet — quantified."),
        ("p", "  - OFI corrected: v1 said 'bid_ask_imbalance mostly NULL' — WRONG. [FACT] the column does not exist at all; code uses flow_imbalance. Claude's error, corrected."),
        ("p", "  - m5_ block: [FACT] it is lagged 1m data, not a 5m timeframe. FINAL DISPOSITION proposed = DROP. (Earlier 'build 5/15/60m bar tables' idea RETRACTED — research says multi-rate pooling inside one model, not separate bar tables.)"),
        ("blank", ""),
        ("h", "OPEN QUESTIONS for Cursor (Claude has NOT resolved these — give your view)"),
        ("p", "  Q1. Independent 12-model design vs joint shared-encoder + per-horizon heads — operator LIKES the 12-independent design. Does emergent coherence via shared features (ladder sheet) suffice, or is enforced coherence (joint model) worth the negative-transfer risk? See Architecture sheet."),
        ("p", "  Q2. Triple-barrier vs fixed-horizon labels — operator's concern: triple-barrier may 'miss a move' by exiting on first barrier. How do we confirm/quantify the missed-opportunity cost? (proposed: dual-label backtest, measure disagreement + realized edge.)"),
        ("p", "  Q3. Monte Carlo — [FACT] traced: underutilized (sizing/narrative/display yes; fusion weight 0; not in ML cone). BAR_MINUTES=5 vs 1m label = blocking for efe/eae-as-feature. Cursor agrees."),
        ("p", "  Q4. Should the 4 horizon cards align? Independent models give NO coherence guarantee; the shared-feature ladder gives EMERGENT (soft) coherence. Is that the right amount, or do we want enforced term-structure?"),
        ("p", "  Q5. Which features should the nets (LSTM/Transformer) gain that they don't have today (delta/oi walls, volume, mega-cap stream)? XGB-only rows are candidates."),
    ]
    r = 1
    for kind, text in lines:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=14, color="1F3864")
        elif kind == "sub":
            cell.font = Font(italic=True, size=10, color="595959"); cell.alignment = WRAP
        elif kind == "h":
            cell.font = WHITE; cell.fill = SUB
        elif kind == "fact":
            cell.fill = FACT; cell.alignment = WRAP
        elif kind == "research":
            cell.fill = RESEARCH; cell.alignment = WRAP
        elif kind == "framing":
            cell.fill = FRAMING; cell.alignment = WRAP
        else:
            cell.alignment = WRAP
        r += 1

    # ===== Sheet 2: FULL STACK x HORIZON (the 6-layer grid) =====
    ws2 = wb.create_sheet("Stack x Horizon (FULL)")
    _stack_grid(ws2)

    # ===== Sheet 3: FEATURE MASTER (canonical, tier-tagged) + Cursor cols =====
    ws3 = wb.create_sheet("Feature Master")
    cols = ["Feature group", "Members", "Source (leaf/derived)", "Pop% SPY/QQQ/IWM",
            "Current consumers", "1c", "5c", "15c", "60c",
            "Research basis [RESEARCH]", "Claude framing [FRAMING] / [FACT]",
            "CURSOR: agree/disagree", "CURSOR: reasoning + sources"]
    for j, c in enumerate(cols, 1):
        _wrapcell(ws3, 1, j, c, fill=(CURSORCOL if c.startswith("CURSOR") else HDR),
                  font=(BOLD if c.startswith("CURSOR") else WHITE))
    widths = [22, 40, 22, 14, 22, 6, 6, 6, 6, 46, 50, 16, 40]
    for j, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for (key, label, members, src, pop, cons, hz, research, framing) in GROUPS:
        _wrapcell(ws3, r, 1, label, fill=GRP, font=BOLD)
        _wrapcell(ws3, r, 2, members)
        _wrapcell(ws3, r, 3, src)
        _wrapcell(ws3, r, 4, pop, align=CTR)
        _wrapcell(ws3, r, 5, cons)
        for k, h in enumerate(HORIZONS):
            st = hz[h]
            _wrapcell(ws3, r, 6 + k, st, fill=STATUS_FILL.get(st), align=CTR)
        _wrapcell(ws3, r, 10, research, fill=RESEARCH)
        _wrapcell(ws3, r, 11, framing, fill=FRAMING)
        cv = CURSOR_VERDICTS.get(key, ("", ""))
        _wrapcell(ws3, r, 12, cv[0], fill=CURSORCOL)
        _wrapcell(ws3, r, 13, cv[1], fill=CURSORCOL)
        ws3.row_dimensions[r].height = 92
        r += 1
    # legend
    r += 1
    _wrapcell(ws3, r, 1, "Legend:", font=BOLD, border=False)
    for k, (lab, fl) in enumerate([("KEEP", KEEP), ("TEST", TEST), ("ADD", ADD), ("DROP", DROP), ("n/a", NA)]):
        _wrapcell(ws3, r, 2 + k, lab, fill=fl, align=CTR)

    # ===== Sheet 4: CROSS-HORIZON LADDER (the operator-validated overlap) =====
    ws4 = wb.create_sheet("Cross-horizon ladder")
    ws4.column_dimensions["A"].width = 3
    ws4.column_dimensions["B"].width = 122
    ls = _ladder_stats()
    ladder = [
        ("title", "CROSS-HORIZON SHARED-FEATURE LADDER  [FACT — computed from Feature Master GROUPS]"),
        ("sub", "Operator insight (2026-06-01): in the 12-independent design, information & coherence emerge from features SHARED between adjacent horizons. Quantified below (recomputed each build)."),
        ("blank", ""),
        ("h", "Shared feature-groups between adjacent horizons (any model uses them at both)"),
    ]
    for line in ls["pair_txt"]:
        ladder.append(("fact", line))
    ladder += [
        ("blank", ""),
        ("h", f"The shared SPINE — {len(ls['spine'])} groups used at ALL FOUR horizons [FACT]"),
        ("p", ", ".join(ls["spine"]) + "."),
        ("h", "FAST-only (present at 1c, gone by 60c) [FACT]"),
        ("p", ", ".join(ls["fast"]) + "."),
        ("h", "SLOW-only (present at 60c, not at 1c) [FACT]"),
        ("p", ", ".join(ls["slow"]) + "."),
        ("blank", ""),
        ("h", "[FRAMING] what this implies (Cursor: agree/disagree)"),
        ("p", "  - The 12-independent models are NOT isolated: neighbours share ~all features, so adjacent cards will tend to MOVE TOGETHER (emergent coherence) without any shared weights."),
        ("p", "  - You should expect a soft term-structure across the 4 cards (spine drives agreement; fast/slow features drive the differences), NOT random alternation. Wild up/down/up/down across horizons = likely low-conviction noise, not insight."),
        ("p", "  - Cards do NOT have to align (1c can differ from 60c legitimately), but high-conviction regimes should show alignment across the spine. This is the 'reading the cards as a term-structure of direction' lens."),
        ("p", "  - OPEN: is EMERGENT coherence (shared features) enough, or do we want ENFORCED coherence (joint model)? Operator prefers the 12-independent design; this ladder is the argument that it already buys most of the coherence. CURSOR: verify + give your verdict."),
        ("blank", ""),
        ("h", "CURSOR notes (fill in):"),
        ("cursor", CURSOR_LADDER),
    ]
    r = 1
    for kind, text in ladder:
        cell = ws4.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=13, color="1F3864")
        elif kind == "sub":
            cell.font = Font(italic=True, size=10, color="595959"); cell.alignment = WRAP
        elif kind == "h":
            cell.font = WHITE; cell.fill = SUB
        elif kind == "fact":
            cell.fill = FACT
        elif kind == "cursor":
            cell.fill = CURSORCOL; ws4.row_dimensions[r].height = 40
        else:
            cell.alignment = WRAP
        r += 1

    # ===== Sheet 5: ARCHITECTURE decision (independent vs joint) =====
    ws5 = wb.create_sheet("Architecture")
    ws5.column_dimensions["A"].width = 3
    ws5.column_dimensions["B"].width = 122
    arch = [
        ("title", "ARCHITECTURE — independent 12-model stack vs joint shared-encoder. CURSOR: independent verdict required."),
        ("blank", ""),
        ("h", "[FACT] current design (verified)"),
        ("p", "12 ML stack layer slots = 3 architectures (XGB, LSTM, Transformer) x 4 horizons (1c/5c/15c/60c), each trained independently. Horizons are FORWARD-LABEL windows over 1m bars (OUTCOME_HORIZON_MINUTES = {1c:1,5c:5,15c:15,60c:60}), NOT separate timeframes. One real bar table: price_bars_1m. Fused via Meta (LogReg over layer probs) + bayesian_fusion (regime+MC+rules)."),
        ("blank", ""),
        ("h", "[RESEARCH] what the literature says (lean deep-research; cite + challenge)"),
        ("research", "  - Joint shared-encoder + per-horizon DIRECT heads (TFT Lim 2021 arXiv:1912.09363; N-HiTS AAAI 2023 arXiv:2201.12886) beats independent on daily/hourly (~7-40%). Intraday-direct evidence (Zhang&Zohren 2021 arXiv:2105.10430): shared encoder comparable short-horizon, SUPERIOR long-horizon (short feeds long)."),
        ("research", "  - BUT all favourable intraday evidence is tick/LOB-sampled, NOT 1-minute clock bars -> magnitude on our substrate is UNPROVEN (extrapolation)."),
        ("research", "  - Negative transfer is real (arXiv:2101.09645): naive hard-sharing across 4 horizons x K tickers can degrade heads -> needs per-horizon loss weighting / gated sharing shipped WITH the joint model."),
        ("research", "  - Multi-rate pooling INSIDE one model (N-HiTS) > separate 5/15/60m bar tables. DIRECT (one pass) > recursive for noise-heavy 1m (arXiv:2511.11461)."),
        ("research", "  - The joint-model win is a NETS property (shared latent). XGBoost has no shared representation to transfer -> stays independent per-horizon. (all models, all timeframes.)"),
        ("blank", ""),
        ("h", "[FRAMING] Claude's hypothesis (challengeable)"),
        ("framing", "  - The joint model is the right CHALLENGER but must BEAT the current 12-independent stack on held-out, purged-CV, per-horizon MCC before it ships. The current stack is the champion baseline (the SPY/QQQ/IWM retrain produces that baseline)."),
        ("framing", "  - Operator prefers the 12-independent design. The cross-horizon ladder (Sheet 4) is the argument that it ALREADY buys most of the coherence via shared features. So: keep 12-independent; treat joint as a measured contender, not a rewrite."),
        ("framing", "  - XGB stays independent per-horizon regardless; nets (LSTM/Transformer) are where a joint encoder would be tested."),
        ("blank", ""),
        ("h", "CURSOR verdict (fill in — agree/disagree + your research):"),
        ("cursor", CURSOR_ARCHITECTURE),
    ]
    r = 1
    for kind, text in arch:
        cell = ws5.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=13, color="1F3864")
        elif kind == "h":
            cell.font = WHITE; cell.fill = SUB
        elif kind == "research":
            cell.fill = RESEARCH; cell.alignment = WRAP
        elif kind == "framing":
            cell.fill = FRAMING; cell.alignment = WRAP
        elif kind == "cursor":
            cell.fill = CURSORCOL; ws5.row_dimensions[r].height = 46
        else:
            cell.alignment = WRAP
        r += 1

    # ===== Sheet 6: LABELING + ABLATION method =====
    ws6 = wb.create_sheet("Labeling & Ablation")
    ws6.column_dimensions["A"].width = 3
    ws6.column_dimensions["B"].width = 122
    method = [
        ("title", "LABELING + FEATURE CURATION/ABLATION method. CURSOR: verify the method meets the standard."),
        ("blank", ""),
        ("h", "[RESEARCH] Labeling"),
        ("research", "  - Triple-barrier (Lopez de Prado): first of profit/stop/time barrier; vertical barrier in bars -> one barrier-set per horizon. Path-dependent = matches how you'd really exit -> more tradeable edge than fixed-horizon."),
        ("research", "  - Keep fixed-horizon as the MEASURED baseline (no clean public TBL>FTH numeric proof exists)."),
        ("research", "  - Flat class need NOT collapse: tune barriers to realized vol; keep flat ~= 1/3 (Minority Collapse is a FINITE-imbalance effect, PNAS). MCC + per-class recall as the collapse detector (not balanced accuracy). Class-weighted loss over SMOTE. Purged+embargoed CV mandatory (overlapping per-horizon labels leak)."),
        ("framing", "  - [FRAMING] operator concern: triple-barrier may MISS a move (exits on first barrier). CONFIRM via dual-label backtest: label same data both ways, measure (a) disagreement rate, (b) what happened next on disagreements, (c) realized edge/PnL each way. Decide on the number, not opinion."),
        ("blank", ""),
        ("h", "[RESEARCH] Curate-THEN-ablate (so ONE pass is meaningful — operator: no train/ablate/train loop)"),
        ("research", "  1. DROP XGBoost built-in Gini/gain importance (inflates continuous features = our OFI/GEX)."),
        ("research", "  2. Spearman-hierarchical cluster -> keep one representative per cluster BEFORE ranking (fixes correlated siblings reading as zero — the m5/wall duplication)."),
        ("research", "  3. Held-out (train-vs-test) permutation = leakage detector (big on train, ~0 on test = leak)."),
        ("research", "  4. Grouped MDA — permute whole families (OFI-family, GEX-family) so the model can't recover via a sibling."),
        ("research", "  5. SHAP-interaction for interaction hypotheses (OFI x regime, dGEX x level) — never thin on solo score."),
        ("research", "  6. (if a joint TFT-style net exists) Variable Selection Network = learned per-horizon feature gate — a free second opinion."),
        ("framing", "  - [FRAMING] this is the curation gate that makes the spreadsheet meaningful: features reach ablation already populated + de-duped + leakage-traced, so ablation (which is meaning-agnostic) isn't fooled by correlation/leakage/dead columns."),
        ("blank", ""),
        ("h", "CURSOR notes (fill in):"),
        ("cursor", CURSOR_LABELING),
    ]
    r = 1
    for kind, text in method:
        cell = ws6.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=13, color="1F3864")
        elif kind == "h":
            cell.font = WHITE; cell.fill = SUB
        elif kind == "research":
            cell.fill = RESEARCH; cell.alignment = WRAP
        elif kind == "framing":
            cell.fill = FRAMING; cell.alignment = WRAP
        elif kind == "cursor":
            cell.fill = CURSORCOL; ws6.row_dimensions[r].height = 40
        else:
            cell.alignment = WRAP
        r += 1

    # ===== Sheet 7: NEW candidates + Research basis =====
    ws7 = wb.create_sheet("NEW candidates & sources")
    ws7.column_dimensions["A"].width = 3
    ws7.column_dimensions["B"].width = 122
    nc = [
        ("title", "NEW feature candidates (research-backed) + source list. CURSOR: add/challenge."),
        ("blank", ""),
        ("h", "Candidates not yet in the cone"),
        ("research", "  [P0] single-level OFI — Schwab L1 derivable; dominant 1-5m signal; targets the 1c collapse. Needs L1->stored column wiring (real new plumbing)."),
        ("research", "  [P0] dGEX — first difference of net_gamma; > GEX level OOS; cheap (diff of existing leaf)."),
        ("research", "  [P0] triple-barrier labeling — the label, not a feature; research #1 lever for flat-collapse."),
        ("research", "  [P2] OI-migration (day-over-day OI concentration delta) — forward-looking dealer positioning; needs prior-day OI store."),
        ("research", "  [P1] multi-rate pooling inside the nets (N-HiTS style) — replaces the m5_ lag; enables 15c/60c at their own scale WITHOUT separate bar tables."),
        ("research", "  [P1] MONTE CARLO efe/eae/containment as ML features + per-card envelope — MC already computes a forward excursion forecast per governed horizon and ALREADY drives sizing/narrative/display (call_engine sizing, prediction_engine narrative, Key Levels) — but is NOT in the ML training cone and NOT a per-card 4-horizon envelope. [BLOCKING PRECONDITION — Cursor catch] BAR_MINUTES=5 means MC forward time is 5x the outcome_Nc label clock (1c sim=5min vs label 1min, etc.); FIX this (BAR_MINUTES=1 outcome-aligned, or explicit minutes->bars map) BEFORE wiring efe/eae as features or they are mislabeled vs the training label. Then: (a) efe/eae/containment as per-horizon ML features; (b) per-card risk envelope (extend existing live-horizon sizing to per-card-horizon, tie to LFE)."),
        ("research", "  [HYGIENE] monte_carlo.DEFAULT_HORIZON=13 is DEAD (the only live caller signals.py:519 always passes horizon_bars=_mc_bars; 13 is a relic of the retired 3c/8c/13c scheme, never reached live). Remove for clarity — harmless but misleading."),
        ("blank", ""),
        ("h", "Sources (lean deep-research, 2026-06-01; 35 verified findings)"),
        ("p", "TFT — Lim et al. 2021, arXiv:1912.09363. N-HiTS — Challu et al. AAAI 2023, arXiv:2201.12886. Intraday multi-horizon — Zhang & Zohren 2021, arXiv:2105.10430. LOB direction decay — TLOB 2025, arXiv:2502.15757. Negative transfer — arXiv:2101.09645. Direct-vs-recursive noise — arXiv:2511.11461. Triple-barrier / Minority Collapse — Lopez de Prado 2018; Fang et al. PNAS PMC8639364. MCC — PLOS One pone.0177678. Grouped/conditional permutation — arXiv:2312.10858. TreeSHAP — arXiv:1802.03888. Curate-then-ablate (Spearman-hierarchical) — sklearn multicollinear example; Molnar IML ch.23."),
        ("p", "[NOTE] All favourable intraday architecture evidence is tick/LOB-sampled, not 1-minute clock bars — the single biggest honesty flag. Treat joint-model magnitude as unproven on our substrate."),
        ("blank", ""),
        ("research", "[FACT] MC efe/eae TRACED + Cursor-corrected (2026-06-01): efe=favorable / eae=adverse excursion (bayesian_fusion.py:83-84). FORWARD forecast per governed horizon. STATUS=UNDERUTILIZED (drives sizing call_engine.py:997-1048/:1751 + narrative + display; NOT in ML features, NOT fusion-weighted, NOT a 4-horizon card envelope). BLOCKING: BAR_MINUTES=5 vs 1m outcome label = 5x wall-clock misalignment (monte_carlo.py:32 vs horizon_outcomes.py:42-46) — fix before efe/eae feature wiring."),
        ("research", "[CREDIT] The 'underutilized' relabel + the BAR_MINUTES misalignment are CURSOR's independent-scrutiny catches (2026-06-01), verified by Claude against source. v2's earlier 'computed-unused/pure waste' was overstated and is corrected."),
        ("cursor", "CURSOR: candidates you'd add / challenge:"),
        ("cursor", ""),
    ]
    r = 1
    for kind, text in nc:
        cell = ws7.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=13, color="1F3864")
        elif kind == "h":
            cell.font = WHITE; cell.fill = SUB
        elif kind == "research":
            cell.fill = RESEARCH; cell.alignment = WRAP
        elif kind == "cursor":
            cell.fill = CURSORCOL; ws7.row_dimensions[r].height = 40
        else:
            cell.alignment = WRAP
        r += 1

    out = "feature_assignment_matrix_v2_DRAFT.xlsx"
    wb.save(out)
    return out


def _stack_grid(ws):
    """6-layer stack x 4 horizons. Each cell = does this stack member operate at this horizon,
    and HOW it consumes features. [FACT] where verified."""
    ws.column_dimensions["A"].width = 22
    # header
    _wrapcell(ws, 1, 1, "Stack layer", fill=HDR, font=WHITE)
    for j, h in enumerate(HORIZONS):
        _wrapcell(ws, 1, 2 + j, h, fill=SUB, font=WHITE, align=CTR)
        ws.column_dimensions[get_column_letter(2 + j)].width = 26
    _wrapcell(ws, 1, 6, "Feature consumption [FACT/FRAMING]", fill=HDR, font=WHITE)
    ws.column_dimensions["F"].width = 60
    _wrapcell(ws, 1, 7, "CURSOR: agree/notes", fill=CURSORCOL, font=BOLD)
    ws.column_dimensions["G"].width = 40

    rows = [
        ("XGBoost", ["active", "active", "active", "active"],
         "[FACT] widest cone (~80 engineered feats incl. ALL categoricals + full Greek/wall set). Independent per-horizon. [FRAMING] stays independent — trees have no shared latent to transfer."),
        ("LSTM", ["active", "active", "active", "active"],
         "[FACT] FEATURES_5M (23) over 60x1m + FEATURES_1M (12) over 20x1m + 6 confluence. Sequential. [FRAMING] candidate for joint shared-encoder."),
        ("Transformer", ["active", "active", "active", "active"],
         "[FACT] consumes LSTM seq encoding + [xgb|lstm] OOF base-prob vectors (cascade). [FRAMING] primary candidate for the joint multi-horizon model (TFT-style)."),
        ("Meta (LogReg)", ["active", "active", "active", "active"],
         "[FACT] consumes ONLY the [xgb|lstm|transformer] layer-prob vector — NO raw features. Its 'features' ARE the ML stack layers."),
        ("Monte Carlo", ["underutilized", "underutilized", "underutilized", "underutilized"],
         "[FACT] simulator, NOT a directional learner. monte_carlo.simulate(horizon_bars=N) projects N bars forward -> expected_move, efe (favorable excursion), eae (adverse excursion), containment (monte_carlo.py:289-293). Runs every tick per governed horizon (signals.py:1134-1149 loop; _mc_bars=horizon_slug_to_mc_bars(_hz) :497; 1c->1,5c->5,15c->15,60c->60 bars). [FACT-CORRECTED by Cursor scrutiny 2026-06-01] status is UNDERUTILIZED, not 'unused': mc_efe/mc_eae DO drive money-path DECISIONS -> position sizing (call_engine.py:997-1048 EAE/stop ratio + EFE-stop floor + EAE gate; :1751 compute_position_size), narrative (prediction_engine.py:1188-1212), display (Key Levels static/index.html:6770-6780), and a post-fusion directional nudge via mc_feature_dict (mc_fusion_adjustment.py @ signals.py:1161-1164). NOT used: (i) fusion posterior weight = 0 (bayesian_fusion.py:776); (ii) NOT in the engineer_features ML training cone; (iii) NOT a simultaneous 4-horizon card envelope (sizing/narrative use LIVE-horizon MC only, default 1c, signals.py:1196-1198). [FACT-BLOCKING — Cursor catch] WALL-CLOCK MISALIGNMENT: BAR_MINUTES=5 (monte_carlo.py:32) so each MC step = 5 MINUTES, but ML outcome horizons are 1-MINUTE clock (horizon_outcomes.py:42-46). Effective MC forward time = bars x 5m: 1c->5min (label 1min), 5c->25min (label 5min), 15c->75min (label 15min), 60c->300min (label 60min) — MC simulates 5x TOO FAR vs the training label. The Key Levels header 'horizon ... 1m snapshot clock' (static/index.html:6768) is WRONG — the engine steps 5m. [FRAMING] MC = magnitude/risk axis, orthogonal to direction. P1 disposition (a) efe/eae/containment as per-horizon ML features + (b) per-card risk envelope tied to LFE — PRECONDITION: fix the 5m-vs-1m misalignment FIRST (BAR_MINUTES=1 or explicit minutes->bars map) or efe/eae features are mislabeled vs outcome_Nc; and decide if sizing stays live-horizon-only or goes per-card-horizon. Extend the existing sizing/narrative/display consumers, NOT greenfield."),
        ("Regime/Rules-Fusion", ["active", "active", "active", "active"],
         "[FACT] bayesian_fusion combines xgb/lstm/transformer/meta + MC + regime + rules into the final read (verified: each referenced 40-72x). This is the actual final combiner. [FRAMING] the true 'full stack' the cards should reflect."),
    ]
    r = 2
    statusfill = {"active": KEEP, "advisory": TEST, "computed-unused": DROP, "underutilized": TEST, "n/a": NA}
    for name, perhz, consume in rows:
        _wrapcell(ws, r, 1, name, fill=GRP, font=BOLD)
        for j, st in enumerate(perhz):
            _wrapcell(ws, r, 2 + j, st, fill=statusfill.get(st), align=CTR)
        _wrapcell(ws, r, 6, consume, fill=FRAMING if "[FRAMING]" in consume else FACT)
        _cursor_note = CURSOR_MC_STACK if name == "Monte Carlo" else ""
        _wrapcell(ws, r, 7, _cursor_note, fill=CURSORCOL)
        ws.row_dimensions[r].height = 74
        r += 1
    # note row
    _wrapcell(ws, r + 1, 1, "[FRAMING] operator principle", font=BOLD, border=False)
    _wrapcell(ws, r + 1, 2, "Each card = the FULL stack fused for that horizon (all 6 layers), not xgb/lstm/transformer-only. 'Not using the full stack at each horizon is cheating ourselves.' Ablation per-horizon confirms each layer is ADDITIVE (independent signal) vs dead weight — full-stack is the starting hypothesis, ablation earns each layer's place.",
              align=WRAP)
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)
    ws.row_dimensions[r + 1].height = 60
    ws.freeze_panes = "B2"


if __name__ == "__main__":
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    ap = argparse.ArgumentParser()
    ap.add_argument("--manifest-only", action="store_true", help="Write feature_ablation_manifest.json only")
    ap.add_argument("--manifest-path", default=str(EXPANDED_ABLATION_MANIFEST_PATH))
    ap.add_argument(
        "--build-schwab-ablation-universe",
        action="store_true",
        help="Categorize all Schwab fields + write expanded ablation manifest (≥2× ML cone)",
    )
    ap.add_argument(
        "--design-matrix",
        action="store_true",
        help="Legacy 7-sheet design workbook only (not the operator ablation list)",
    )
    args = ap.parse_args()
    if args.design_matrix:
        print("wrote:", build())
    elif args.build_schwab_ablation_universe or args.manifest_only:
        reg = build_schwab_ablation_field_registry(write=True)
        p = write_feature_ablation_manifest(Path(args.manifest_path))
        totals = resolve_ablation_universe()["totals"]
        print(
            f"wrote registry ({reg['schwab_field_count']} fields) + manifest {p} + "
            f"{FEATURE_ABLATION_UNIVERSE_XLSX}  "
            f"ablation_groups={totals['ablation_group_count']}  "
            f"expansion_ratio={totals.get('expansion_ratio_vs_cone')}"
        )
    else:
        mp = write_feature_ablation_manifest()
        print(f"wrote: {mp} + {FEATURE_ABLATION_UNIVERSE_XLSX}")
